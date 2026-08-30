from __future__ import annotations

import contextlib
import copy
import ctypes
import gc
import html
import os
import re
import shutil
import sys
import threading
import time
from difflib import SequenceMatcher
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Callable

from network_guard import block_python_network


ProgressCallback = Callable[[str, int | None, str, int | None], None]
SUPPORTED_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff", ".pdf"}
RECOGNITION_MODES = {"text", "table"}
MODEL_PROFILES = {
    "fast": {
        "label": "PP-OCRv5 轻量",
        "detection": "PP-OCRv5_mobile_det",
        "recognition": "PP-OCRv5_mobile_rec",
    },
    "accurate": {
        "label": "PP-OCRv5 高精度",
        "detection": "PP-OCRv5_server_det",
        "recognition": "PP-OCRv5_server_rec",
    },
}
TABLE_MODEL_NAMES = ("PicoDet_layout_1x_table", "SLANet_plus")


def model_names(profile: str, mode: str) -> list[str]:
    if profile not in MODEL_PROFILES:
        raise ValueError(f"未知模型档位：{profile}")
    if mode not in RECOGNITION_MODES:
        raise ValueError(f"未知识别模式：{mode}")
    model = MODEL_PROFILES[profile]
    names = [model["detection"], model["recognition"]]
    if mode == "table":
        names = [*TABLE_MODEL_NAMES, *names]
    return names


def official_model_cache() -> Path:
    return Path.home() / ".paddlex" / "official_models"


def _tree_stats(path: Path) -> tuple[int, int]:
    size = 0
    count = 0
    if not path.is_dir() or path.is_symlink():
        return size, count
    for item in path.rglob("*"):
        if item.is_file() and not item.is_symlink():
            try:
                size += item.stat().st_size
                count += 1
            except OSError:
                continue
    return size, count


def model_cache_status(
    profile: str,
    mode: str,
    cache_root: Path | None = None,
) -> dict[str, Any]:
    root = (cache_root or official_model_cache()).expanduser().resolve()
    entries: list[dict[str, Any]] = []
    for name in model_names(profile, mode):
        model_path = root / name
        size, file_count = _tree_stats(model_path)
        def nonempty(filename: str) -> bool:
            candidate = model_path / filename
            try:
                return candidate.is_file() and not candidate.is_symlink() and candidate.stat().st_size > 0
            except OSError:
                return False
        installed = not model_path.is_symlink() and nonempty("inference.pdiparams") and (
            nonempty("inference.json") or nonempty("inference.pdmodel")
        ) and nonempty("inference.yml")
        entries.append(
            {
                "name": name,
                "installed": installed,
                "state": "ready" if installed else "incomplete" if file_count else "missing",
                "sizeBytes": size,
                "fileCount": file_count,
            }
        )
    return {
        "cacheRoot": str(root),
        "profile": profile,
        "mode": mode,
        "installed": all(item["installed"] for item in entries),
        "sizeBytes": sum(int(item["sizeBytes"]) for item in entries),
        "modelCount": len(entries),
        "installedCount": sum(bool(item["installed"]) for item in entries),
        "models": entries,
    }


class ModelProgressStream:
    def __init__(self, stream: Any, names: list[str], callback: ProgressCallback | None):
        self.stream, self.names, self.callback = stream, names, callback
        # .buffer is the underlying binary stream, not our text accumulator.
        self._pending_text = ""
        self.last_name = ""

    def write(self, value: str) -> int:
        result = self.stream.write(value)
        self.stream.flush()
        self._pending_text = (self._pending_text + value)[-2000:]
        for name in self.names:
            if name in self._pending_text and name != self.last_name:
                self.last_name = name
                self._pending_text = ""
                if self.callback:
                    self.callback(f"正在下载或载入模型：{name}", None, "model", None)
                break
        return result

    def __getattr__(self, name: str) -> Any:
        return getattr(self.stream, name)


class TableHtmlParser(HTMLParser):
    """Convert model-generated table HTML to a safe, renderable cell grid."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[dict[str, Any]]] = []
        self._row = -1
        self._column = 0
        self._occupied: set[tuple[int, int]] = set()
        self._cell: dict[str, Any] | None = None
        self._text_parts: list[str] = []

    @staticmethod
    def _span(attributes: dict[str, str | None], name: str) -> int:
        try:
            return max(1, min(int(attributes.get(name) or 1), 100))
        except (TypeError, ValueError):
            return 1

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        name = tag.lower()
        if name == "tr":
            self._row += 1
            self._column = 0
            self.rows.append([])
            return
        if name not in {"td", "th"} or self._row < 0 or self._cell is not None:
            return

        attributes = dict(attrs)
        while (self._row, self._column) in self._occupied:
            self._column += 1
        row_span = self._span(attributes, "rowspan")
        col_span = self._span(attributes, "colspan")
        self._cell = {
            "row": self._row,
            "column": self._column,
            "rowSpan": row_span,
            "colSpan": col_span,
            "text": "",
            "box": [],
        }
        self._text_parts = []
        for row in range(self._row, self._row + row_span):
            for column in range(self._column, self._column + col_span):
                self._occupied.add((row, column))

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._text_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() not in {"td", "th"} or self._cell is None:
            return
        self._cell["text"] = " ".join("".join(self._text_parts).split())
        self.rows[self._row].append(self._cell)
        self._column = int(self._cell["column"]) + int(self._cell["colSpan"])
        self._cell = None
        self._text_parts = []


def parse_table_html(value: str, cell_boxes: list[Any] | None = None) -> list[list[dict[str, Any]]]:
    parser = TableHtmlParser()
    parser.feed(value or "")
    parser.close()
    boxes = list(cell_boxes or [])
    cells = [cell for row in parser.rows for cell in row]
    for index, cell in enumerate(cells):
        if index < len(boxes):
            cell["box"] = _jsonable(boxes[index])
    return [row for row in parser.rows if row]


def safe_table_html(rows: list[list[dict[str, Any]]]) -> str:
    body: list[str] = []
    for row in rows:
        cells: list[str] = []
        for cell in row:
            row_span = max(1, int(cell.get("rowSpan") or 1))
            col_span = max(1, int(cell.get("colSpan") or 1))
            attributes = ""
            if row_span > 1:
                attributes += f' rowspan="{row_span}"'
            if col_span > 1:
                attributes += f' colspan="{col_span}"'
            cells.append(f"<td{attributes}>{html.escape(str(cell.get('text') or ''))}</td>")
        body.append(f"<tr>{''.join(cells)}</tr>")
    return f"<table><tbody>{''.join(body)}</tbody></table>"


def _jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        return {str(key): _jsonable(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonable(item) for item in value]
    if hasattr(value, "tolist"):
        return _jsonable(value.tolist())
    return str(value)


def extract_ocr_payload(payload: dict[str, Any], page_index: int | None = None) -> dict[str, Any]:
    texts = list(payload.get("rec_texts") or [])
    scores = list(payload.get("rec_scores") or [])
    polygons = list(payload.get("rec_polys") or [])
    boxes = list(payload.get("rec_boxes") or [])

    blocks: list[dict[str, Any]] = []
    for index, text in enumerate(texts):
        clean_text = str(text).strip()
        if not clean_text:
            continue
        blocks.append(
            {
                "text": clean_text,
                "score": float(scores[index]) if index < len(scores) else 0.0,
                "polygon": polygons[index] if index < len(polygons) else [],
                "box": boxes[index] if index < len(boxes) else [],
            }
        )

    return {
        "pageIndex": payload.get("page_index", page_index),
        "text": "\n".join(block["text"] for block in blocks),
        "blocks": blocks,
        "tables": [],
    }


def extract_page(result: Any) -> dict[str, Any]:
    """Convert PaddleOCR 3.x's Result.json payload into the stable app schema."""

    raw = _jsonable(result.json)
    payload = raw.get("res", raw) if isinstance(raw, dict) else {}
    return extract_ocr_payload(payload)


def _table_box(cell_boxes: list[Any]) -> list[float]:
    points: list[tuple[float, float]] = []
    for box in cell_boxes:
        values = _jsonable(box)
        if not isinstance(values, list):
            continue
        if values and isinstance(values[0], list):
            for point in values:
                if isinstance(point, list) and len(point) >= 2:
                    points.append((float(point[0]), float(point[1])))
        else:
            numbers = [float(item) for item in values]
            points.extend(zip(numbers[0::2], numbers[1::2]))
    if not points:
        return []
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    return [min(xs), min(ys), max(xs), max(ys)]


def extract_table_page(result: Any) -> dict[str, Any]:
    """Convert TableRecognitionPipelineV2 output to text plus safe table grids."""

    raw = _jsonable(result.json)
    payload = raw.get("res", raw) if isinstance(raw, dict) else {}
    page_index = payload.get("page_index")
    overall_ocr = payload.get("overall_ocr_res") or {}
    page = extract_ocr_payload(overall_ocr, page_index)
    page["pageIndex"] = page_index

    layout_tables = [
        item
        for item in (payload.get("layout_det_res") or {}).get("boxes", [])
        if str(item.get("label") or "").lower() == "table"
    ]
    tables: list[dict[str, Any]] = []
    for index, table_payload in enumerate(payload.get("table_res_list") or []):
        cell_boxes = list(table_payload.get("cell_box_list") or [])
        rows = parse_table_html(str(table_payload.get("pred_html") or ""), cell_boxes)
        layout = layout_tables[index] if index < len(layout_tables) else {}
        box = _table_box(cell_boxes) or list(layout.get("coordinate") or [])
        tables.append(
            {
                "pageIndex": page_index,
                "endPageIndex": page_index,
                "tableIndex": index,
                "sourceTableCount": 1,
                "score": float(layout["score"]) if layout.get("score") is not None else None,
                "box": box,
                "html": safe_table_html(rows),
                "rows": rows,
            }
        )
    page["tables"] = tables
    return page


def _table_column_count(table: dict[str, Any]) -> int:
    return max(
        (
            int(cell.get("column") or 0) + max(1, int(cell.get("colSpan") or 1))
            for row in table.get("rows") or []
            for cell in row
        ),
        default=0,
    )


def _normalized_cell_text(value: Any) -> str:
    return re.sub(r"[\W_]+", "", str(value or "").casefold(), flags=re.UNICODE)


def _rows_match(left: list[dict[str, Any]], right: list[dict[str, Any]]) -> bool:
    if len(left) != len(right) or not left:
        return False
    left_structure = [
        (int(cell.get("column") or 0), max(1, int(cell.get("colSpan") or 1)))
        for cell in left
    ]
    right_structure = [
        (int(cell.get("column") or 0), max(1, int(cell.get("colSpan") or 1)))
        for cell in right
    ]
    if left_structure != right_structure:
        return False

    pairs = [
        (_normalized_cell_text(left_cell.get("text")), _normalized_cell_text(right_cell.get("text")))
        for left_cell, right_cell in zip(left, right)
    ]
    if not any(left_text and right_text for left_text, right_text in pairs):
        return False
    similarities = [
        SequenceMatcher(None, left_text, right_text).ratio()
        if left_text and right_text
        else float(left_text == right_text)
        for left_text, right_text in pairs
    ]
    return sum(similarities) / len(similarities) >= 0.82 and sum(
        similarity >= 0.72 for similarity in similarities
    ) >= max(1, (3 * len(similarities) + 3) // 4)


def _repeated_header_rows(left: dict[str, Any], right: dict[str, Any]) -> int:
    if _table_column_count(left) != _table_column_count(right):
        return 0
    left_rows = list(left.get("rows") or [])
    right_rows = list(right.get("rows") or [])
    repeated = 0
    for index in range(min(3, len(left_rows), len(right_rows))):
        if not _rows_match(left_rows[index], right_rows[index]):
            break
        repeated += 1
    return repeated


def merge_cross_page_tables(pages: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Conservatively merge one-table pages that repeat the same leading header."""

    merged: list[dict[str, Any]] = []
    previous_page_tables: list[dict[str, Any]] = []
    for page in pages:
        page_tables = list(page.get("tables") or [])
        page_index = page.get("pageIndex")
        if len(page_tables) == 1 and len(previous_page_tables) == 1 and merged:
            candidate = merged[-1]
            candidate_end = candidate.get("endPageIndex")
            if (
                isinstance(page_index, int)
                and isinstance(candidate_end, int)
                and page_index == candidate_end + 1
            ):
                repeated_rows = _repeated_header_rows(candidate, page_tables[0])
                if repeated_rows:
                    continuation_rows = copy.deepcopy(page_tables[0].get("rows") or [])[repeated_rows:]
                    existing_end = max(
                        (
                            int(cell.get("row") or 0) + max(1, int(cell.get("rowSpan") or 1))
                            for row in candidate.get("rows") or []
                            for cell in row
                        ),
                        default=0,
                    )
                    continuation_start = min(
                        (
                            int(cell.get("row") or 0)
                            for row in continuation_rows
                            for cell in row
                        ),
                        default=existing_end,
                    )
                    row_offset = existing_end - continuation_start
                    for row in continuation_rows:
                        for cell in row:
                            cell["row"] = int(cell.get("row") or 0) + row_offset
                    candidate["rows"].extend(continuation_rows)
                    candidate["endPageIndex"] = page_index
                    candidate["sourceTableCount"] = int(candidate.get("sourceTableCount") or 1) + 1
                    candidate["html"] = safe_table_html(candidate["rows"])
                    previous_page_tables = page_tables
                    continue

        for table in page_tables:
            normalized = copy.deepcopy(table)
            normalized["endPageIndex"] = normalized.get("pageIndex")
            normalized["sourceTableCount"] = 1
            merged.append(normalized)
        previous_page_tables = page_tables
    return merged


def document_page_count(path: Path) -> int:
    if path.suffix.lower() != ".pdf":
        return 1
    from pypdfium2 import PdfDocument

    document = PdfDocument(str(path))
    try:
        return len(document)
    finally:
        document.close()


class OcrEngine:
    def __init__(self) -> None:
        self._ocr: Any | None = None
        self._profile: str | None = None
        self._mode: str | None = None
        self._native_runtime: Any | None = None
        self._runtime_initialized = False
        self._pause_requested = threading.Event()
        self._cancel_requested = threading.Event()

    @property
    def ready(self) -> bool:
        return self._ocr is not None

    @property
    def profile(self) -> str | None:
        return self._profile

    @property
    def mode(self) -> str | None:
        return self._mode

    def initialize_runtime(self, progress: ProgressCallback | None = None) -> None:
        """Cold-import native dependencies before the command loop resumes stdin.

        0.7.0 moved first imports into the prepare worker as well as model loading.
        Keep first-time imports on the main thread; model creation/downloads can
        still run in the worker after imports finish. Do not wrap stderr during
        cold imports: third-party logging setup may inspect its binary interface.
        """
        if self._runtime_initialized:
            return
        if threading.current_thread() is not threading.main_thread():
            raise RuntimeError("Paddle 依赖首次导入必须由 sidecar 主线程执行")
        with contextlib.redirect_stdout(sys.stderr):
            if progress:
                progress("正在主线程导入 Paddle（尚未创建模型）……", None, "import_paddle", None)
            import paddle  # noqa: F401
            if progress:
                progress("Paddle 导入完成，正在导入 PaddleOCR / PaddleX……", None, "import_paddleocr", None)
            from paddleocr import PaddleOCR, TableRecognitionPipelineV2  # noqa: F401
        self._runtime_initialized = True
        if progress:
            progress("OCR 依赖导入完成，即将创建模型流水线", None, "imports_ready", None)

    def check_native_runtime(self) -> dict[str, Any]:
        """Verify that Paddle's lazily loaded CPU runtime is discoverable."""

        if os.name == "nt":
            library_name = "mklml.dll"
            loader = ctypes.WinDLL
        elif sys.platform.startswith("linux"):
            library_name = "libmklml_intel.so"
            loader = ctypes.CDLL
        else:
            raise RuntimeError(f"Unsupported sidecar runtime platform: {sys.platform}")

        # Keep the handle alive for the rest of the process so Paddle can reuse
        # the already loaded library when it creates a predictor.
        if self._native_runtime is None:
            self._native_runtime = loader(library_name)
        return {"ok": True, "library": library_name}

    def unload(self) -> None:
        self._ocr = None
        self._profile = None
        self._mode = None
        gc.collect()

    def delete_model_cache(self, profile: str, mode: str) -> dict[str, Any]:
        root = official_model_cache().expanduser().resolve()
        before = model_cache_status(profile, mode, root)
        self.unload()
        removed: list[str] = []
        for name in model_names(profile, mode):
            target = root / name
            if not target.exists():
                continue
            if target.is_symlink() or target.resolve().parent != root:
                raise RuntimeError(f"拒绝删除非官方模型缓存目录：{target}")
            if not target.is_dir():
                raise RuntimeError(f"模型缓存目标不是目录：{target}")
            shutil.rmtree(target)
            removed.append(name)
        return {
            "removed": removed,
            "freedBytes": before["sizeBytes"],
            "status": model_cache_status(profile, mode, root),
        }

    def prepare(
        self,
        profile: str = "fast",
        mode: str = "text",
        progress: ProgressCallback | None = None,
    ) -> dict[str, Any]:
        if profile not in MODEL_PROFILES:
            raise ValueError(f"未知模型档位：{profile}")
        if mode not in RECOGNITION_MODES:
            raise ValueError(f"未知识别模式：{mode}")
        model = MODEL_PROFILES[profile]
        mode_label = "轻量表格与文字" if mode == "table" else model["label"]
        if self._ocr is not None and self._profile == profile and self._mode == mode:
            return {
                "ready": True,
                "downloaded": False,
                "model": mode_label,
                "profile": profile,
                "mode": mode,
            }

        if progress:
            progress(f"正在检查并下载{mode_label}模型……", None, "status", None)

        if self._ocr is not None:
            self.unload()

        # PaddleOCR may emit progress information. Keep stdout reserved for NDJSON.
        progress_stream = ModelProgressStream(sys.stderr, model_names(profile, mode), progress)
        with contextlib.redirect_stdout(progress_stream), contextlib.redirect_stderr(progress_stream):
            if mode == "table":
                from paddleocr import TableRecognitionPipelineV2
                if progress:
                    progress("正在创建轻量表格流水线；需要时下载模型……", None, "create_pipeline", None)
                self._ocr = TableRecognitionPipelineV2(
                    layout_detection_model_name="PicoDet_layout_1x_table",
                    wired_table_structure_recognition_model_name="SLANet_plus",
                    wireless_table_structure_recognition_model_name="SLANet_plus",
                    text_detection_model_name=model["detection"],
                    text_recognition_model_name=model["recognition"],
                    use_doc_orientation_classify=False,
                    use_doc_unwarping=False,
                    use_layout_detection=True,
                    use_ocr_model=True,
                    device="cpu",
                    cpu_threads=max(1, min(os.cpu_count() or 4, 8)),
                )
            else:
                from paddleocr import PaddleOCR
                if progress:
                    progress("正在创建文字识别流水线；需要时下载模型……", None, "create_pipeline", None)
                self._ocr = PaddleOCR(
                    text_detection_model_name=model["detection"],
                    text_recognition_model_name=model["recognition"],
                    use_doc_orientation_classify=False,
                    use_doc_unwarping=False,
                    use_textline_orientation=False,
                    device="cpu",
                    cpu_threads=max(1, min(os.cpu_count() or 4, 8)),
                )
            self._profile = profile
            self._mode = mode

        if progress:
            progress(f"{mode_label}模型已下载并载入本机内存", None, "status", None)
        return {
            "ready": True,
            "downloaded": True,
            "model": mode_label,
            "profile": profile,
            "mode": mode,
        }

    def reset_job_control(self) -> None:
        self._pause_requested.clear()
        self._cancel_requested.clear()

    def pause(self) -> None:
        self._pause_requested.set()

    def resume(self) -> None:
        self._pause_requested.clear()

    def cancel(self) -> None:
        self._cancel_requested.set()
        self._pause_requested.clear()

    def _wait_if_paused(self, page: int, page_count: int, progress: ProgressCallback | None) -> None:
        if not self._pause_requested.is_set() or self._cancel_requested.is_set():
            return
        if progress:
            progress(f"识别已暂停（第 {page}/{page_count} 页）；已完成的结果会保留", page, "paused", page_count)
        while self._pause_requested.is_set() and not self._cancel_requested.wait(0.1):
            pass
        if progress and not self._cancel_requested.is_set():
            progress(f"继续本地识别，第 {page + 1}/{page_count} 页……", page, "resumed", page_count)

    def recognize(
        self,
        path_value: str,
        score_threshold: float = 0.5,
        mode: str = "text",
        progress: ProgressCallback | None = None,
        on_page: Callable[[dict[str, Any], int, int], None] | None = None,
    ) -> dict[str, Any]:
        path = Path(path_value).expanduser().resolve(strict=True)
        if not path.is_file():
            raise ValueError("所选路径不是文件")
        if path.suffix.lower() not in SUPPORTED_SUFFIXES:
            raise ValueError(f"不支持的文件类型：{path.suffix}")
        if not 0.0 <= score_threshold <= 1.0:
            raise ValueError("最低置信度必须在 0 到 1 之间")
        if self._ocr is None:
            raise RuntimeError("模型尚未准备，请先执行 prepare")
        if mode != self._mode:
            raise RuntimeError("识别模式与已载入模型不一致，请重新准备模型")

        started = time.perf_counter()
        pages: list[dict[str, Any]] = []
        total_page_count = document_page_count(path)
        if progress:
            if path.suffix.lower() == ".pdf":
                prefix = "开始定位表格、识别文字并恢复结构" if mode == "table" else "开始识别"
                message = f"已读取 PDF，共 {total_page_count} 页；已封锁 Python 网络连接，{prefix}……"
            else:
                message = (
                    "已封锁 Python 网络连接，开始定位表格、识别文字并恢复结构……"
                    if mode == "table"
                    else "已封锁 Python 网络连接，开始读取本地图片……"
                )
            progress(message, 0, "status", total_page_count)

        with block_python_network(), contextlib.redirect_stdout(sys.stderr):
            predict_options: dict[str, Any] = {
                "input": str(path),
                "text_rec_score_thresh": score_threshold,
            }
            if mode == "table":
                # TableRecognitionPipelineV2 defaults this separate orientation
                # switch to True at prediction time. PaddleX then lazily creates
                # and downloads its orientation classifier even though document
                # orientation was disabled when the pipeline was constructed.
                # Recognition runs behind the network guard, so the lightweight
                # profile must explicitly disable every optional preprocessing
                # model here as well as in prepare().
                predict_options.update(
                    {
                        "use_doc_orientation_classify": False,
                        "use_doc_unwarping": False,
                        "use_layout_detection": True,
                        "use_ocr_model": True,
                        "use_table_orientation_classify": False,
                    }
                )
            results = self._ocr.predict_iter(**predict_options)
            for page_number, result in enumerate(results, start=1):
                page = extract_table_page(result) if mode == "table" else extract_page(result)
                if page.get("pageIndex") is None:
                    page["pageIndex"] = page_number - 1
                for table in page["tables"]:
                    table["pageIndex"] = page["pageIndex"]
                    table["endPageIndex"] = page["pageIndex"]
                pages.append(page)
                if on_page:
                    on_page(page, round((time.perf_counter() - started) * 1000), total_page_count)
                if progress:
                    percent = round(page_number / total_page_count * 100)
                    detail = (
                        f"，识别到 {len(page['tables'])} 个表格"
                        if mode == "table"
                        else ""
                    )
                    progress(
                        f"已完成第 {page_number}/{total_page_count} 页（{percent}%）{detail}",
                        page_number,
                        "progress",
                        total_page_count,
                    )
                if self._cancel_requested.is_set():
                    break
                self._wait_if_paused(page_number, total_page_count, progress)
                if self._cancel_requested.is_set():
                    break

        elapsed_ms = round((time.perf_counter() - started) * 1000)
        tables = merge_cross_page_tables(pages) if mode == "table" else []
        return {
            "path": str(path),
            "profile": self._profile,
            "resultType": mode,
            "cancelled": self._cancel_requested.is_set(),
            "text": "\n\n".join(page["text"] for page in pages if page["text"]),
            "pageCount": len(pages),
            "totalPageCount": total_page_count,
            "blockCount": sum(len(page["blocks"]) for page in pages),
            "tableCount": len(tables),
            "rawTableCount": sum(len(page["tables"]) for page in pages),
            "elapsedMs": elapsed_ms,
            "pages": pages,
            "tables": tables,
        }


def export_text_results(directory_value: str, items: list[dict[str, Any]]) -> dict[str, Any]:
    directory = Path(directory_value).expanduser().resolve(strict=True)
    if not directory.is_dir():
        raise ValueError("导出位置不是文件夹")

    exported: list[dict[str, str]] = []
    reserved: set[Path] = set()
    for item in items:
        source_name = Path(str(item.get("fileName") or "识别结果")).name
        stem = Path(source_name).stem.strip() or "识别结果"
        candidate = directory / f"{stem}.txt"
        index = 2
        while candidate.exists() or candidate in reserved:
            candidate = directory / f"{stem} ({index}).txt"
            index += 1
        candidate.write_text(str(item.get("text") or ""), encoding="utf-8")
        reserved.add(candidate)
        exported.append({"source": source_name, "path": str(candidate)})
    return {"count": len(exported), "files": exported}


def _available_path(directory: Path, name: str, reserved: set[Path]) -> Path:
    requested = Path(name)
    candidate = directory / requested.name
    index = 2
    while candidate.exists() or candidate in reserved:
        candidate = directory / f"{requested.stem} ({index}){requested.suffix}"
        index += 1
    reserved.add(candidate)
    return candidate


def export_table_results(
    directory_value: str,
    items: list[dict[str, Any]],
    formats: list[str] | None = None,
) -> dict[str, Any]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    directory = Path(directory_value).expanduser().resolve(strict=True)
    if not directory.is_dir():
        raise ValueError("导出位置不是文件夹")
    requested_formats = set(formats if formats is not None else ["xlsx", "html"])
    unsupported_formats = requested_formats - {"xlsx", "html"}
    if unsupported_formats:
        raise ValueError(f"不支持的表格导出格式：{', '.join(sorted(unsupported_formats))}")
    if not requested_formats:
        raise ValueError("至少选择一种表格导出格式")

    exported: list[dict[str, Any]] = []
    reserved: set[Path] = set()
    for item in items:
        tables = list(item.get("tables") or [])
        if not tables:
            continue
        source_name = Path(str(item.get("fileName") or "表格识别结果")).name
        stem = Path(source_name).stem.strip() or "表格识别结果"
        xlsx_path = (
            _available_path(directory, f"{stem}.xlsx", reserved)
            if "xlsx" in requested_formats
            else None
        )
        html_path = (
            _available_path(directory, f"{stem}.tables.html", reserved)
            if "html" in requested_formats
            else None
        )

        workbook = Workbook() if xlsx_path is not None else None
        if workbook is not None:
            workbook.remove(workbook.active)
        html_sections: list[str] = []
        for ordinal, table in enumerate(tables, start=1):
            page_number = int(table.get("pageIndex") or 0) + 1
            end_page_number = int(table.get("endPageIndex") or table.get("pageIndex") or 0) + 1
            page_label = (
                f"第 {page_number}–{end_page_number} 页"
                if end_page_number > page_number
                else f"第 {page_number} 页"
            )
            table_number = int(table.get("tableIndex") or (ordinal - 1)) + 1
            page_token = (
                f"P{page_number}-{end_page_number}"
                if end_page_number > page_number
                else f"P{page_number}"
            )
            base_title = f"{page_token}-T{table_number}"[:31]
            source_label = str(table.get("sourceName") or "")
            if source_label:
                short_source = re.sub(r"[\\/\[\]:*?]", "_", source_label)[:16]
                base_title = f"{short_source}-{base_title}"[:31]
            title = base_title
            suffix = 2
            if workbook is not None:
                while title in workbook.sheetnames:
                    marker = f"-{suffix}"
                    title = f"{base_title[:31 - len(marker)]}{marker}"
                    suffix += 1
                sheet = workbook.create_sheet(title)
            else:
                sheet = None

            rows = list(table.get("rows") or [])
            if sheet is not None:
                for row in rows:
                    for cell in row:
                        row_index = int(cell.get("row") or 0) + 1
                        column_index = int(cell.get("column") or 0) + 1
                        row_span = max(1, int(cell.get("rowSpan") or 1))
                        col_span = max(1, int(cell.get("colSpan") or 1))
                        target = sheet.cell(row=row_index, column=column_index)
                        target.value = str(cell.get("text") or "")
                        # OCR text is data, never an Excel formula.
                        target.data_type = "s"
                        target.alignment = Alignment(vertical="center", wrap_text=True)
                        if row_index == 1:
                            target.font = Font(bold=True)
                            target.fill = PatternFill("solid", fgColor="E7EFE9")
                        if row_span > 1 or col_span > 1:
                            sheet.merge_cells(
                                start_row=row_index,
                                start_column=column_index,
                                end_row=row_index + row_span - 1,
                                end_column=column_index + col_span - 1,
                            )
                        width = min(max(len(str(target.value)) + 2, 10), 40)
                        column_letter = target.column_letter
                        sheet.column_dimensions[column_letter].width = max(
                            sheet.column_dimensions[column_letter].width or 0,
                            width / col_span,
                        )

            if html_path is not None:
                table_html = safe_table_html(rows)
                html_sections.append(
                    f"<section><h2>{html.escape(source_label)} {page_label} · 表格 {table_number}</h2>{table_html}</section>"
                )

        file_result: dict[str, Any] = {"source": source_name, "tableCount": len(tables)}
        if workbook is not None and xlsx_path is not None:
            workbook.save(xlsx_path)
            file_result["xlsx"] = str(xlsx_path)
        if html_path is not None:
            html_document = (
                "<!doctype html><html lang=\"zh-CN\"><head><meta charset=\"utf-8\">"
                f"<title>{html.escape(stem)} 表格识别结果</title>"
                "<style>body{font-family:sans-serif;margin:24px;color:#18201d}"
                "section{margin-bottom:32px}table{border-collapse:collapse;max-width:100%;}"
                "td{border:1px solid #68716b;padding:6px 9px;vertical-align:top}"
                "h1{font-size:22px}h2{font-size:16px}</style></head><body>"
                f"<h1>{html.escape(source_name)}</h1>{''.join(html_sections)}</body></html>"
            )
            html_path.write_text(html_document, encoding="utf-8")
            file_result["html"] = str(html_path)
        exported.append(file_result)

    return {
        "count": len(exported),
        "tableCount": sum(int(item["tableCount"]) for item in exported),
        "formatCounts": {
            output_format: sum(output_format in item for item in exported)
            for output_format in sorted(requested_formats)
        },
        "files": exported,
    }
