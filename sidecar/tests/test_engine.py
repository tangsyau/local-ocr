from __future__ import annotations

import sys
import io
import os
import threading
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

SIDECAR_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SIDECAR_DIR))

from engine import (  # noqa: E402
    MODEL_PROFILES,
    ModelProgressStream,
    OcrEngine,
    document_page_count,
    export_table_results,
    export_text_results,
    extract_page,
    extract_table_page,
    merge_cross_page_tables,
    model_cache_status,
    model_names,
    parse_table_html,
    safe_table_html,
)
from network_guard import NetworkBlockedError, block_python_network  # noqa: E402


class FakeArray:
    def __init__(self, value: object) -> None:
        self.value = value

    def tolist(self) -> object:
        return self.value


class FakeResult:
    json = {
        "res": {
            "page_index": 2,
            "rec_texts": ["第一行", "", "second line"],
            "rec_scores": FakeArray([0.98, 0.2, 0.75]),
            "rec_polys": FakeArray([[[0, 0], [10, 0], [10, 5], [0, 5]], [], [[0, 8], [20, 8], [20, 14], [0, 14]]]),
            "rec_boxes": FakeArray([[0, 0, 10, 5], [], [0, 8, 20, 14]]),
        }
    }


class FakeTableResult:
    json = {
        "res": {
            "page_index": 1,
            "overall_ocr_res": {
                "rec_texts": ["部门", "姓名", "技术部", "张三"],
                "rec_scores": FakeArray([0.99, 0.98, 0.97, 0.96]),
                "rec_polys": FakeArray([[], [], [], []]),
                "rec_boxes": FakeArray([[0, 0, 40, 20], [40, 0, 80, 20], [0, 20, 40, 40], [40, 20, 80, 40]]),
            },
            "layout_det_res": {
                "boxes": [
                    {"label": "table", "score": 0.91, "coordinate": [0, 0, 80, 40]}
                ]
            },
            "table_res_list": [
                {
                    "cell_box_list": FakeArray(
                        [[0, 0, 40, 0, 40, 20, 0, 20], [40, 0, 80, 0, 80, 20, 40, 20], [0, 20, 40, 20, 40, 40, 0, 40]]
                    ),
                    "pred_html": (
                        "<html><body><table><tr><td rowspan='2'>部门</td><td>姓名</td></tr>"
                        "<tr><td><script>alert(1)</script>张三</td></tr></table></body></html>"
                    ),
                }
            ],
        }
    }


class EngineSchemaTests(unittest.TestCase):
    def test_progress_stream_preserves_binary_buffer_and_flushes_model_events(self) -> None:
        # Default text streams translate LF to the platform line ending. Exercise
        # explicit LF and CRLF on every OS too, so Linux tests cover Windows bytes.
        for newline, expected_ending in ((None, os.linesep), ("\n", "\n"), ("\r\n", "\r\n")):
            with self.subTest(newline=newline):
                binary = io.BytesIO()
                with io.TextIOWrapper(binary, encoding="utf-8", newline=newline) as stream:
                    events = []
                    wrapped = ModelProgressStream(stream, ["model-A"], lambda *args: events.append(args))
                    self.assertIs(wrapped.buffer, binary)
                    wrapped.write("Creating model: model-")
                    wrapped.write("A\n")
                    # Check before close/flush: the wrapper must flush immediately.
                    self.assertEqual(binary.getvalue(), ("Creating model: model-A" + expected_ending).encode("utf-8"))
                    self.assertEqual(len(events), 1)
                    self.assertEqual(events[0][2], "model")

    def test_runtime_cold_imports_run_once_on_main_thread(self) -> None:
        engine = OcrEngine()
        events = []
        original_stdout = sys.stdout
        with patch.dict(sys.modules, {
            "paddle": SimpleNamespace(),
            "paddleocr": SimpleNamespace(PaddleOCR=object, TableRecognitionPipelineV2=object),
        }), patch("engine.initialize_document_runtime") as initialize_documents:
            engine.initialize_runtime(lambda *args: events.append(args))
            engine.initialize_runtime(lambda *args: events.append(args))
        initialize_documents.assert_called_once()
        self.assertEqual([entry[2] for entry in events], ["import_paddle", "import_paddleocr", "import_documents", "imports_ready"])
        self.assertIs(sys.stdout, original_stdout)

    def test_runtime_cold_imports_are_rejected_in_worker(self) -> None:
        errors = []
        def attempt() -> None:
            try:
                OcrEngine().initialize_runtime()
            except RuntimeError as error:
                errors.append(str(error))
        worker = threading.Thread(target=attempt)
        worker.start()
        worker.join(timeout=2)
        self.assertFalse(worker.is_alive())
        self.assertEqual(len(errors), 1)
        self.assertIn("主线程", errors[0])

    def test_extract_page_returns_stable_schema(self) -> None:
        page = extract_page(FakeResult())
        self.assertEqual(page["pageIndex"], 2)
        self.assertEqual(page["text"], "第一行\nsecond line")
        self.assertEqual(len(page["blocks"]), 2)
        self.assertAlmostEqual(page["blocks"][1]["score"], 0.75)
        self.assertEqual(page["blocks"][0]["box"], [0, 0, 10, 5])
        self.assertEqual(page["tables"], [])

    def test_table_html_is_parsed_to_safe_spanning_cells(self) -> None:
        rows = parse_table_html(
            "<table><tr><td rowspan='2'>部门</td><td>姓名</td></tr>"
            "<tr><td><img src=x onerror=alert(1)>张三</td></tr></table>"
        )
        self.assertEqual(rows[0][0]["rowSpan"], 2)
        self.assertEqual(rows[1][0]["column"], 1)
        self.assertEqual(rows[1][0]["text"], "张三")
        rendered = safe_table_html(rows)
        self.assertNotIn("<img", rendered)
        self.assertIn('rowspan="2"', rendered)

    def test_extract_table_page_returns_text_and_structured_tables(self) -> None:
        page = extract_table_page(FakeTableResult())
        self.assertEqual(page["pageIndex"], 1)
        self.assertIn("部门", page["text"])
        self.assertEqual(len(page["tables"]), 1)
        table = page["tables"][0]
        self.assertAlmostEqual(table["score"], 0.91)
        self.assertEqual(table["box"], [0.0, 0.0, 80.0, 40.0])
        self.assertNotIn("<script", table["html"])

    def test_python_network_is_denied_inside_guard(self) -> None:
        import socket

        with block_python_network():
            with self.assertRaises(NetworkBlockedError):
                socket.create_connection(("example.com", 80))

    def test_model_profiles_use_mobile_and_server_pairs(self) -> None:
        self.assertEqual(MODEL_PROFILES["fast"]["detection"], "PP-OCRv5_mobile_det")
        self.assertEqual(MODEL_PROFILES["accurate"]["recognition"], "PP-OCRv5_server_rec")

    def test_model_cache_status_is_limited_to_known_models(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            cache_root = Path(temp_dir)
            model_dir = cache_root / "PP-OCRv5_mobile_det"
            model_dir.mkdir()
            (model_dir / "inference.json").write_bytes(b"model")
            (model_dir / "inference.pdiparams").write_bytes(b"params")
            (model_dir / "inference.yml").write_bytes(b"config")
            status = model_cache_status("fast", "text", cache_root)

        self.assertEqual(model_names("fast", "text"), ["PP-OCRv5_mobile_det", "PP-OCRv5_mobile_rec"])
        self.assertEqual(status["installedCount"], 1)
        self.assertFalse(status["installed"])
        self.assertEqual(status["sizeBytes"], 17)

    def test_table_model_manifest_reuses_one_structure_model(self) -> None:
        names = model_names("fast", "table")
        self.assertEqual(names[:2], ["PicoDet_layout_1x_table", "SLANet_plus"])
        self.assertEqual(len(names), len(set(names)))

    def test_delete_model_cache_keeps_unrelated_directories(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            cache_root = Path(temp_dir)
            for name in [*model_names("fast", "text"), "unrelated-model"]:
                model_dir = cache_root / name
                model_dir.mkdir()
                (model_dir / "model.bin").write_bytes(b"1234")
            engine = OcrEngine()
            engine._ocr = object()
            engine._profile = "fast"
            engine._mode = "text"
            with patch("engine.official_model_cache", return_value=cache_root):
                result = engine.delete_model_cache("fast", "text")

            self.assertEqual(set(result["removed"]), set(model_names("fast", "text")))
            self.assertTrue((cache_root / "unrelated-model").is_dir())
            self.assertFalse(engine.ready)

    def test_table_mode_uses_lightweight_structure_pipeline(self) -> None:
        captured: dict[str, object] = {}

        class FakeTablePipeline:
            def __init__(self, **kwargs: object) -> None:
                captured.update(kwargs)

        fake_module = SimpleNamespace(TableRecognitionPipelineV2=FakeTablePipeline)
        with patch.dict(sys.modules, {"paddleocr": fake_module}):
            engine = OcrEngine()
            result = engine.prepare("fast", "table")

        self.assertEqual(result["mode"], "table")
        self.assertEqual(captured["layout_detection_model_name"], "PicoDet_layout_1x_table")
        self.assertEqual(captured["wired_table_structure_recognition_model_name"], "SLANet_plus")
        self.assertEqual(captured["wireless_table_structure_recognition_model_name"], "SLANet_plus")
        self.assertEqual(captured["text_detection_model_name"], "PP-OCRv5_mobile_det")
        self.assertFalse(captured["use_doc_orientation_classify"])

    def test_table_recognition_disables_lazy_orientation_downloads(self) -> None:
        captured: dict[str, object] = {}

        class FakeTablePipeline:
            def predict_iter(self, **kwargs: object) -> object:
                captured.update(kwargs)
                return iter([FakeTableResult()])

        with tempfile.TemporaryDirectory() as temp_dir:
            image_path = Path(temp_dir) / "table.png"
            from PIL import Image
            with Image.new("RGB", (80, 40), "white") as image:
                image.save(image_path)
            engine = OcrEngine()
            engine._ocr = FakeTablePipeline()
            engine._profile = "fast"
            engine._mode = "table"
            result = engine.recognize(str(image_path), mode="table")

        self.assertEqual(result["resultType"], "table")
        self.assertFalse(captured["use_doc_orientation_classify"])
        self.assertFalse(captured["use_doc_unwarping"])
        self.assertFalse(captured["use_table_orientation_classify"])
        self.assertTrue(captured["use_layout_detection"])
        self.assertTrue(captured["use_ocr_model"])

    def test_repeated_headers_merge_adjacent_single_table_pages(self) -> None:
        def table(page_index: int, header: str, value: str) -> dict[str, object]:
            rows = [
                [
                    {"row": 0, "column": 0, "rowSpan": 1, "colSpan": 1, "text": header, "box": []},
                    {"row": 0, "column": 1, "rowSpan": 1, "colSpan": 1, "text": "金额", "box": []},
                ],
                [
                    {"row": 1, "column": 0, "rowSpan": 1, "colSpan": 1, "text": value, "box": []},
                    {"row": 1, "column": 1, "rowSpan": 1, "colSpan": 1, "text": "100", "box": []},
                ],
            ]
            return {
                "pageIndex": page_index,
                "endPageIndex": page_index,
                "tableIndex": 0,
                "sourceTableCount": 1,
                "score": 0.9,
                "box": [],
                "html": safe_table_html(rows),
                "rows": rows,
            }

        merged = merge_cross_page_tables(
            [
                {"pageIndex": 0, "tables": [table(0, "项目", "第一项")]},
                {"pageIndex": 1, "tables": [table(1, "项目", "第二项")]},
                {"pageIndex": 2, "tables": [table(2, "项目", "第三项")]},
            ]
        )
        self.assertEqual(len(merged), 1)
        self.assertEqual(merged[0]["endPageIndex"], 2)
        self.assertEqual(merged[0]["sourceTableCount"], 3)
        self.assertEqual(len(merged[0]["rows"]), 4)
        self.assertEqual(merged[0]["rows"][-1][0]["row"], 3)
        self.assertEqual(merged[0]["rows"][-1][0]["text"], "第三项")

        not_merged = merge_cross_page_tables(
            [
                {"pageIndex": 0, "tables": [table(0, "项目", "第一项")]},
                {"pageIndex": 1, "tables": [table(1, "姓名", "张三")]},
            ]
        )
        self.assertEqual(len(not_merged), 2)

    def test_export_text_results_avoids_name_collisions(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            first = Path(temp_dir) / "报告.txt"
            first.write_text("existing", encoding="utf-8")
            result = export_text_results(
                temp_dir,
                [
                    {"fileName": "报告.pdf", "text": "第一份"},
                    {"fileName": "报告.png", "text": "第二份"},
                ],
            )
            exported = [Path(item["path"]) for item in result["files"]]
            self.assertEqual([path.name for path in exported], ["报告 (2).txt", "报告 (3).txt"])
            self.assertEqual(exported[0].read_text(encoding="utf-8"), "第一份")

    def test_export_table_results_writes_xlsx_and_safe_html(self) -> None:
        from openpyxl import load_workbook

        page = extract_table_page(FakeTableResult())
        with tempfile.TemporaryDirectory() as temp_dir:
            result = export_table_results(
                temp_dir,
                [{"fileName": "名单.pdf", "tables": page["tables"]}],
            )
            self.assertEqual(result["count"], 1)
            self.assertEqual(result["tableCount"], 1)
            xlsx_path = Path(result["files"][0]["xlsx"])
            html_path = Path(result["files"][0]["html"])
            workbook = load_workbook(xlsx_path)
            sheet = workbook[workbook.sheetnames[0]]
            self.assertEqual(sheet["A1"].value, "部门")
            self.assertEqual(sheet["B2"].value, "alert(1)张三")
            self.assertIn("A1:A2", {str(item) for item in sheet.merged_cells.ranges})
            html_text = html_path.read_text(encoding="utf-8")
            self.assertNotIn("<script", html_text)

    def test_export_table_results_writes_only_selected_formats(self) -> None:
        page = extract_table_page(FakeTableResult())
        with tempfile.TemporaryDirectory() as temp_dir:
            html_result = export_table_results(
                temp_dir,
                [{"fileName": "名单.pdf", "tables": page["tables"]}],
                ["html"],
            )
            exported = html_result["files"][0]
            self.assertNotIn("xlsx", exported)
            self.assertIn("html", exported)
            self.assertEqual(html_result["formatCounts"], {"html": 1})
            self.assertEqual(list(Path(temp_dir).glob("*.xlsx")), [])
        with tempfile.TemporaryDirectory() as temp_dir:
            xlsx_result = export_table_results(
                temp_dir,
                [{"fileName": "名单.pdf", "tables": page["tables"]}],
                ["xlsx"],
            )
            exported = xlsx_result["files"][0]
            self.assertIn("xlsx", exported)
            self.assertNotIn("html", exported)
            self.assertEqual(xlsx_result["formatCounts"], {"xlsx": 1})
            self.assertEqual(list(Path(temp_dir).glob("*.html")), [])

    def test_export_table_results_rejects_empty_or_unknown_formats(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "至少选择"):
                export_table_results(temp_dir, [], [])
            with self.assertRaisesRegex(ValueError, "不支持"):
                export_table_results(temp_dir, [], ["csv"])

    def test_pdf_page_count_uses_pdfium_and_closes_document(self) -> None:
        closed = False

        class FakeDocument:
            def __init__(self, path: str) -> None:
                self.path = path

            def __len__(self) -> int:
                return 12

            def close(self) -> None:
                nonlocal closed
                closed = True

        with patch.dict(sys.modules, {"pypdfium2": SimpleNamespace(PdfDocument=FakeDocument)}):
            self.assertEqual(document_page_count(Path("example.pdf")), 12)
        self.assertTrue(closed)


if __name__ == "__main__":
    unittest.main()
